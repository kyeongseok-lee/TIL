from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "ExcelSheet"

# A1 셀에 1 이라는 값을 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) # A1 셀의 정보를 출력
print(ws["A1"].value) # A1 셀의 '값'을 출력
print(ws["A10"].value) # 값이 없을 땐 'None' 출력

# row = 1, 2, 3, ...
# column = A, B, C, .. 
print(ws.cell(row=1, column=1).value) # ws["A1"].value
print(ws.cell(row=1, column=2).value) # ws["B1"].value

c = ws.cell(row=1, column=3, value=10) # ws["C1"].value = 10
print(c.value) # ws["C1"].value

from random import *

# 반복을 이용해서 랜덤 숫자 채우기
index = 1
for x in range(1, 11): # 10개 row
    for y in range(1, 11): # 10개 column
        # ws.cell(row=x, column=y, value=randint(0, 100)) # 0~100 사이의 숫자
        ws.cell(row=x, column=y, value=index)
        index += 1
wb.save("sample.xlsx")